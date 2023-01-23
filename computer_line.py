import openpyxl
import re


def get_path():
    path = input('输入文件路径：')
    path = path.split('\"')[1]
    workbook = openpyxl.load_workbook(path)
    return workbook,1,path

def not_empty(s):
    return s and s.strip()

def is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return 1

    return 0

def getdata(file):
    number = input('输入你要提取的格子数量和第一个格子的编号，格式为:1,C4 \n记得要大写,并在英文模式输入\n')
    sheet = file.active
    num = number.split(',')[0]
    location = number.split(',')[1]
    datas = []
    final = []
    cu = sheet.cell(2, 1).value
    cu = cu[-5:]
    cu = int(cu)
    cu = cu // 1000
    cu *= 1000
    for i in range(int(num)):
        data = sheet[location].value
        datas.append(data)
        raw = re.sub(r"[A-Za-z]","",location)
        raw = int(raw)
        raw += 1
        location = re.sub(r"[\d]","",location)+str(raw)
    for val in datas:
        val = re.split(',|mm2| |/|-',val)
        val = filter(not_empty, val)
        val = list(val)
        print(val)
        describe = []
        data = []
        guige = ''
        for x in val:
            x = re.sub(r'X','×',x)
            x = re.sub(r'x','×',x)
            # print('x',x)
            if x.find('×') == -1 and x[0] != '0'and x[0] != '1'and x[0] != '2'and x[0] != '3'and x[0] != '4'and x[0] != '5'and x[0] != '6'and x[0] != '7'and x[0] != '8'and x[0] != '9' or x.find('22')!=-1 or x.find('23')!=-1 or x.find('32')!=-1 or x.find('33')!=-1:
                if x.isdigit() == False or len(x) <= 4:
                    describe.append(x)
        # for m in describe:
        #     for r in m:
        #         if
        print(describe)
        flag = 0
        lab = ''
        for y in describe:
            flag+=is_chinese(y)
        if flag == len(describe):
            lab = 'all_chinese'
        else:
            lab = 'have_english'
        print(lab)
        if lab == 'have_english':  #有英文
            english = ''
            for x in describe:
                if is_chinese(x) == 0:
                    english += x
            chinese = max(describe,key=len)
            print('English',english)
            print('chinese',chinese)
            k = ''
            if english.find('22') != -1 or chinese.find('22') != -1:
                k = '22'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('23') != -1 or chinese.find('23') != -1:
                k = '23'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('32') != -1 or chinese.find('32') != -1:
                k = '32'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('33') != -1 or chinese.find('33') != -1:
                k = '33'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('92') != -1 or chinese.find('92') != -1:
                k = '92'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('93') != -1 or chinese.find('93') != -1:
                k ='93'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            else:
                k = '0'


            if english.find('ZA') != -1 or english.find('ZRA') != -1:
                english = re.sub('ZA', '', english)
                english = re.sub('ZRA', '', english)
                data.append('A')
            elif english.find('ZB') != -1 or english.find('ZRB') != -1:
                english = re.sub('ZB', '', english)
                english = re.sub('ZRA', '', english)
                data.append('B')
            elif english.find('ZC') != -1 or english.find('ZRC') != -1 or english.find('ZR') != -1:
                english = re.sub('ZC', '', english)
                english = re.sub('ZRC', '', english)
                english = re.sub('ZR', '', english)
                data.append('C')
            else:
                data.append('no_zuran')

            #以上是判断阻燃
            if english.find('R') != -1:
                english = re.sub('R','',english)
                data.append('R')
            elif english.find('B') != -1:
                english = re.sub('B','',english)
                data.append('B')
            else:
                data.append('no_soft_heart')

            #以上是判断软心
            if english.find('YP') != -1:
                if english.find('YP2') != -1:
                    data.append('fenping2')
                    english = re.sub('YP2','',english)
                elif english.find('YP3') != -1:
                    data.append('fenping3')
                    english = re.sub('YP3','',english)
                else:
                    english = re.sub('YP','',english)
                    data.append('fenping1')
            else:
                data.append('no_fenping')

            #以上是判断分屏
            if english.find('P') != -1:
                if english.find('P2') != -1:
                    english = re.sub('P2','',english)
                    data.append('zongping2')
                elif english.find('P3') != -1:
                    english = re.sub('P3', '', english)
                    data.append('zongping3')
                else:
                    english = re.sub('P', '', english)
                    data.append('zongping')
            else:
                data.append('no_zongping')

            #以上是判断总屏
            if chinese.find('耐火') != -1 or english.find('NH') != -1 or english.find('N') != -1:
                data.append('anti_fire')
            else:
                data.append('fire_right')
            #以上是耐火
            if chinese.find('白蚁') != -1:
                data.append('anti_ants')
            else:
                data.append('ant_right')
            #以上是防白蚁
            if chinese.find('鼠') != -1:
                data.append('anti_mouse')
            else:
                data.append('mouse_right')
            #以上是防鼠
            if chinese.find('低温') != -1:
                data.append('anti_cold')
            else:
                data.append('cold_right')
            #以上是耐低温
            if chinese.find('无卤') != -1:
                data.append('no_smoke')
            elif chinese.find('低卤') != -1:
                data.append('low_smoke')
            else:
                data.append('smoke_right')
            #以上是低烟低卤
            if english.find('IA') != -1 or chinese.find('本安') != -1:
                data.append('ben_an')
            else:
                data.append('no_ben_an')
            #以上是本安
            if k == '22':
                data.append('22')
            elif k == '23':
                data.append('23')
            elif k == '32':
                data.append('32')
            elif k == '33':
                data.append('33')
            elif k == '92':
                data.append('92')
            elif k == '93':
                data.append('93')
            else:
                data.append('no_kaizhuang')
            #以上是判断铠装
            #阻燃，软心，分屏，总屏，耐火，白蚁，鼠，低温，低烟无卤，本安，铠装
        else:#没英文，用中文
            data = []
            print('aaaaa')
            chinese = max(describe,key=len)
            print(chinese)
            if chinese.find('A') != -1:
                data.append("A")
            elif chinese.find('B') != -1:
                data.append("B")
            elif chinese.find('C') != -1 or chinese.find('ZR') != -1:
                data.append("C")
            else:
                data.append('no_zuran')
            #以上是判断阻燃
            if chinese.find('软') != -1 or chinese.find('R') != -1 or chinese.find('多股') != -1:
                data.append('R')
            elif chinese.find('B') != -1:
                data.append('B')
            else:
                data.append('none')
            #以上是软心
            p = 0
            if chinese.find('编织') != -1:
                p = 1
            elif chinese.find('铜带') != -1:
                p = 2
            elif chinese.find('铝塑') != -1:
                p = 3
            # print(p)
            #以上是判断p类型
            if chinese.find('分屏') != -1 or chinese.find('对屏') != -1:
                data.append('fenping'+str(p))
            else:
                data.append('no_fenping')
            #以上是分屏
            if chinese.find('总屏') != -1:
                data.append('zongping'+str(p))
            else:
                data.append('no_zongping')
            #以上是总屏
            if chinese.find('耐火') != -1:
                data.append('anti_fire')
            else:
                data.append('fire_right')
            #以上是耐火
            if chinese.find('白蚁') != -1:
                data.append('anti_ants')
            else:
                data.append('ant_right')
                # 以上是防白蚁
            if chinese.find('鼠') != -1:
                data.append('anti_mouse')
            else:
                data.append('mouse_right')
                # 以上是防鼠
            if chinese.find('低温') != -1:
                data.append('anti_cold')
            else:
                data.append('cold_right')
                # 以上是耐低温
            if chinese.find('无卤') != -1:
                data.append('no_smoke')
            elif chinese.find('低卤') != -1:
                data.append('low_smoke')
            else:
                data.append('smoke_right')
                # 以上是低烟低卤
            if  chinese.find('本安') != -1:
                data.append('ben_an')
            else:
                data.append('no_ben_an')
                # 以上是本安
            if  chinese.find('22') != -1 and chinese.find('mm23') == -1:
                data.append('22')
            elif  chinese.find('23') != -1 and chinese.find('mm23') == -1:
                data.append('23')
            elif  chinese.find('32') != -1:
                data.append('32')
            elif chinese.find('33') != -1:
                data.append('33')
            elif  chinese.find('92') != -1:
                data.append('92')
            elif  chinese.find('93') != -1:
                data.append('93')
            else:
                data.append('no_kaizhuang')
        for x in val:
            if x.find('×') != -1 or x.find('x') != -1 or x.find('X') != -1:
                x = re.sub(r'mm2', '', x)
                x = re.sub(r'X', '×', x)
                x = re.sub(r'x', '×', x)
                guige = x
                break
        data.append(guige)
        final.append(data)
        temp = []
        for i in final:
            t = []
            for m in i:
                x = re.sub(' ','',m)
                t.append(x)
            temp.append(t)
        final = temp
    return final,cu

def datacreat(x):
    final = []
    for data in x:
        flag = 0
        guige = data[-1]
        print('guige',guige)
        guige = guige.split('×')
        if guige[0] == 1:
            flag = 1
        temp = []
        if len(data) != 12:
            temp = ['error']
        else:
            print(data[-1][0:2])
            if data[-1][0:2] == '1×':
                temp.append('DJYVP')
            else:
                if flag == 1:
                    temp.append('DJYVP')
                if data[2] == 'no_fenping':
                    if data[3] != 'no_zongping':
                        temp.append('DJYVP')
                else:
                    if data[3] == 'no_zongping':
                        temp.append('DJYPV')
                    else:
                        temp.append('DJYPVP')
            if data[2] == 'fenping2' or data[3] == 'zongping2':
                temp.append('p2')
            elif data[2] =='fenping3' or data[3] == 'zongping3':
                temp.append('p3')
            else:
                temp.append('p1')
            temp.append(data[0])
            temp.append(data[1])
            temp.append(data[4])
            temp.append(data[5])
            temp.append(data[6])
            temp.append(data[7])
            temp.append(data[8])
            temp.append(data[9])
            temp.append(data[10])
            temp.append(data[11])
        final.append(temp)
    return final



def getprice(base,data,cu,file,_):
    c = []
    line_of_price = []
    final_list = []
    if cu > 25000 and cu <= 30000:
        line_of_price = [5,6]
        c = [25000,30000]
    elif cu > 30000 and cu <= 35000:
        line_of_price = [6,7]
        c = [30000,35000]
    elif cu > 35000 and cu <= 40000:
        line_of_price = [7,8]
        c =[35000,40000]
    elif cu > 40000 and cu <= 45000:
        line_of_price = [8,9]
        c = [40000,45000]
    elif cu > 45000 and cu <= 50000:
        line_of_price = [9,10]
        c = [45000,50000]
    elif cu > 50000 and cu <= 55000:
        line_of_price = [10,11]
        c = [50000,55000]
    elif cu > 55000 and cu <= 60000:
        line_of_price = [11,12]
        c = [55000,60000]
    elif cu > 60000 and cu <= 65000:
        line_of_price = [12,13]
        c = [60000,65000]
    elif cu > 65000 and cu <= 70000:
        line_of_price = [13,14]
        c = [65000,70000]
    elif cu > 70000 and cu <= 75000:
        line_of_price = [14,15]
        c = [70000,75000]
    elif cu > 75000 and cu <= 80000:
        line_of_price = [15,16]
        c = [75000,80000]
    sheet = base.active
    # print(data)
    for d in data:
        if d == ['error']:
            final_list.append('error')
            continue
        price_list = []
        guige = [d[0],d[-1]]
        # print(guige)
        h = 0
        for m in range(sheet.min_row,sheet.max_row+1):
            h += 1
            cell = sheet.cell(m,4).value
            # print(cell)
            if cell == guige[1]:
                if sheet.cell(h,2).value == guige[0]:
                    price1 = sheet.cell(h,line_of_price[0]+1).value
                    price2 = sheet.cell(h,line_of_price[1]+1).value
                    priceplus = 0.0
                    if d[4] == 'anti_fire':
                        priceplus += sheet.cell(h, 18).value
                    if d[2] == 'A':
                        priceplus += sheet.cell(h, 19).value
                    elif d[2] == 'B':
                        priceplus += sheet.cell(h, 20).value
                    elif d[2] == 'C':
                        priceplus += sheet.cell(h, 21).value
                    if d[8] == 'no_smoke':
                        priceplus += sheet.cell(h, 22).value
                    elif d[8] == 'no_smoke':
                        priceplus += sheet.cell(h, 36).value
                    if d[-2] == '22':
                        priceplus += sheet.cell(h, 26).value
                    elif d[-2] == '23':
                        priceplus += sheet.cell(h, 27).value
                    elif d[-2] == '32':
                        priceplus += sheet.cell(h, 28).value
                    elif d[-2] == '33':
                        priceplus += sheet.cell(h, 29).value
                    elif d[-2] == '92':
                        priceplus += sheet.cell(h, 30).value
                    elif d[-2] == '93':
                        priceplus += sheet.cell(h, 31).value
                    if d[5] == 'anti_ant':
                        priceplus += sheet.cell(h, 32).value
                    if d[7] == 'anti_cold':
                        priceplus += sheet.cell(h, 33).value
                    if d[1] == 'p1':
                        priceplus += 0
                    elif d[1] == 'p2':
                        priceplus += sheet.cell(h, 34).value
                    elif d[1] == 'p3':
                        priceplus += sheet.cell(h, 35).value
                    if d[3] == 'R':
                        priceplus += sheet.cell(h, 25).value
                    elif d[3] == 'B':
                        priceplus += sheet.cell(h, 24).value
                    if d[6] == 'anti_mouse':
                        priceplus += sheet.cell(h, 37).value
                    if d[9] == 'ben_an':
                        priceplus += sheet.cell(h, 23).value
                    price_list.append(h)
                    price_list.append(price1)
                    price_list.append(price2)
                    the = cu - c[0]
                    price = (price2 - price1) / 5000 * the + price1 + priceplus
                    price_list.append(price)
                    final_list.append(price_list)
                    break
        if price_list == []:
            price_list = 'error'
            final_list.append(price_list)
    return final_list

def toexcel(pricelist,file,_,path):
    sheet = file.active
    n = sheet.max_column
    for i in range(len(pricelist)):
        # print(pricelist[i])
        if pricelist[i] != 'error':
            sheet.cell(i+4, n + 1, pricelist[i][0])
            sheet.cell(i+4, n + 2, pricelist[i][1])
            sheet.cell(i+4, n + 3, pricelist[i][2])
            sheet.cell(i+4, n + 4, pricelist[i][3])
        else:
            sheet.cell(i + 4, n + 1, '出错')
            sheet.cell(i + 4, n + 2, '出错')
            sheet.cell(i + 4, n + 3, '出错')
            sheet.cell(i + 4, n + 4, '出错')
        file.save(filename=path)



def computer():
    file, _, path = get_path()
    data, cu = getdata(file)
    print(data)
    print('输入基础表地址')
    data = datacreat(data)
    print(data)
    base, a, path1 = get_path()
    x = getprice(base, data, cu, file, _)
    print(x)
    toexcel(x, file, _, path)
    print('写入完成')



