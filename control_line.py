import openpyxl
import re


def get_path():
    path = input('输入文件路径：')
    path = path.split('\"')[1]
    workbook = openpyxl.load_workbook(path)
    return workbook,1,path

def not_empty(s):
    return s and s.strip()

def is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return 1

    return 0

def getdata(file):
    number = input('输入你要提取的格子数量和第一个格子的编号，格式为:1,C4 \n记得要大写,并在英文模式输入\n')
    sheet = file.active
    num = number.split(',')[0]
    location = number.split(',')[1]
    datas = []
    final = []
    cu = sheet.cell(2, 1).value
    cu = cu[-5:]
    cu = int(cu)
    cu = cu // 1000
    cu *= 1000
    for i in range(int(num)):
        data = sheet[location].value
        datas.append(data)
        raw = re.sub(r"[A-Za-z]","",location)
        raw = int(raw)
        raw += 1
        location = re.sub(r"[\d]","",location)+str(raw)
    for val in datas:
        val = re.split(',|mm2| |/|-',val)
        val = filter(not_empty, val)
        val = list(val)
        print(val)
        describe = []
        data = []
        guige = ''
        for x in val:
            x = re.sub(r'X','×',x)
            x = re.sub(r'x','×',x)
            # print('x',x)

            if x.find('×') == -1 and x[0] != '0'and x[0] != '1'and x[0] != '2'and x[0] != '3'and x[0] != '4'and x[0] != '5'and x[0] != '6'and x[0] != '7'and x[0] != '8'and x[0] != '9' or x.find('22')!=-1 or x.find('23')!=-1 or x.find('32')!=-1 or x.find('33')!=-1:
                if x.isdigit()==False or len(x)<=4:
                    describe.append(x)
        # for m in describe:
        #     for r in m:
        #         if
        print(describe)
        flag = 0
        lab = ''
        for y in describe:
            flag+=is_chinese(y)
        if flag == len(describe):
            lab = 'all_chinese'
        else:
            lab = 'have_english'
        print(lab)
        if lab == 'have_english':  #有英文
            english = ''
            chinese = ''
            for x in describe:
                if is_chinese(x) == 0:
                    english += x
            for x in describe:
                if is_chinese(x) == 1:
                    chinese += x
            print('English',english)
            print('chinese',chinese)
            k = ''
            if english.find('YJV'):
                english = re.sub('yjV', '', english)
                data.append('YJV')
            elif english.find('YJY'):
                english = re.sub('yjV', '', english)
                data.append('YJY')

            if english.find('22') != -1 or chinese.find('22') != -1:
                k = '22'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('23') != -1 or chinese.find('23') != -1:
                k = '23'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('32') != -1 or chinese.find('32') != -1:
                k = '32'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            elif english.find('33') != -1 or chinese.find('33') != -1:
                k = '33'
                english = re.sub(k,'',english)
                chinese = re.sub(k,'',chinese)
            else:
                k = '0'

            if english.find('ZA') != -1 or english.find('ZRA') != -1 or chinese.find('ZA') != -1 or chinese.find('ZRA') != -1:
                english = re.sub('ZA', '', english)
                english = re.sub('ZRA', '', english)
                chinese = re.sub('ZA', '', chinese)
                chinese = re.sub('ZRA', '', chinese)
                data.append('A')
            elif english.find('ZB') != -1 or english.find('ZRB') != -1 or chinese.find('ZB') != -1 or chinese.find('ZRB') != -1:
                english = re.sub('ZB', '', english)
                english = re.sub('ZRB', '', english)
                chinese = re.sub('ZB', '', chinese)
                chinese = re.sub('ZRB', '', chinese)
                data.append('B')
            elif english.find('ZC') != -1 or english.find('ZRC') != -1 or chinese.find('ZC') != -1 or chinese.find('ZRC') != -1:
                english = re.sub('ZC', '', english)
                english = re.sub('ZRC', '', english)
                chinese = re.sub('ZC', '', chinese)
                chinese = re.sub('ZRC', '', chinese)
                data.append('C')
            else:
                data.append('no_zuran')

            #以上是判断阻燃

            if english.find('R') != -1 or chinese.find('软') != -1:
                data.append('soft_heart')
            else:
                data.append('solid_heart')

            if english.find('P') != -1:
                if english.find('P2') != -1 and english.find('P22') == -1 and english.find('P23') == -1:
                    english = re.sub('P2','',english)
                    data.append('p2')
                elif english.find('P3') != -1 and english.find('P32') == -1 and english.find('P33') == -1:
                    english = re.sub('P3', '', english)
                    data.append('p3')
                elif english.find('P4') != -1:
                    english = re.sub('P4', '', english)
                    data.append('p3')
                else:
                    english = re.sub('P', '', english)
                    data.append('p1')
            else:
                data.append('none')

            #以上是判断总屏
            if chinese.find('耐火') != -1 or english.find('NH') != -1 or english.find('N') != -1:
                data.append('anti_fire')
            else:
                data.append('fire_right')
            #以上是耐火
            if chinese.find('白蚁') != -1:
                data.append('anti_ants')
            else:
                data.append('ant_right')
            #以上是防白蚁
            if chinese.find('鼠') != -1:
                data.append('anti_mouse')
            else:
                data.append('mouse_right')
            #以上是防鼠
            if chinese.find('低温') != -1:
                data.append('anti_cold')
            else:
                data.append('cold_right')
            #以上是耐低温
            if chinese.find('无卤') != -1:
                data.append('no_smoke')
            elif chinese.find('低卤') != -1:
                data.append('low_smoke')
            else:
                data.append('smoke_right')
            #以上是低烟低卤
            if english.find('IA') != -1 or chinese.find('本安') != -1:
                data.append('ben_an')
            else:
                data.append('no_ben_an')
            #以上是本安
            if k == '22':
                data.append('22')
            elif k == '23':
                data.append('23')
            elif k == '32':
                data.append('32')
            elif k == '33':
                data.append('33')
            else:
                data.append('no_kaizhuang')
            #以上是判断铠装
            #阻燃，软心，分屏，总屏，耐火，白蚁，鼠，低温，低烟无卤，本安，铠装
        else:#没英文，用中文
            data = []
            print('aaaaa')
            chinese = max(describe,key=len)
            print(chinese)

            if chinese.find('聚乙'):
                data.append('YJY')
            elif chinese.find('聚氯乙'):
                data.append('YJV')
            #以上是判断型号
            if chinese.find('A') != -1:
                data.append("A")
            elif chinese.find('B') != -1:
                data.append("B")
            elif chinese.find('C') != -1 or chinese.find('ZR') != -1:
                data.append("C")
            else:
                data.append('no_zuran')
            #以上是判断阻燃
            if chinese.find('软') != -1:
                data.append('soft_heart')
            else:
                data.append('solid_heart')
            #以上是判断软心
            if chinese.find('编织') != -1 or chinese.find('铜丝') != -1:
                p = 'p1'
            elif chinese.find('铜带') != -1:
                p = 'p2'
            elif chinese.find('铝塑') != -1:
                p = 'p3'
            elif chinese.find('铜塑') != -1:
                p = 'p4'
            else:
                p = 'none'
            data.append(p)
            # print(p)
            #以上是判断p类型
            if chinese.find('耐火') != -1:
                data.append('anti_fire')
            else:
                data.append('fire_right')
            #以上是耐火
            if chinese.find('白蚁') != -1:
                data.append('anti_ants')
            else:
                data.append('ant_right')
                # 以上是防白蚁
            if chinese.find('鼠') != -1:
                data.append('anti_mouse')
            else:
                data.append('mouse_right')
                # 以上是防鼠
            if chinese.find('低温') != -1:
                data.append('anti_cold')
            else:
                data.append('cold_right')
                # 以上是耐低温
            if chinese.find('无卤') != -1:
                data.append('no_smoke')
            elif chinese.find('低卤') != -1:
                data.append('low_smoke')
            else:
                data.append('smoke_right')
                # 以上是低烟低卤
            if  chinese.find('22') != -1 and chinese.find('mm23') == -1:
                data.append('22')
            elif  chinese.find('23') != -1 and chinese.find('mm23') == -1:
                data.append('23')
            elif  chinese.find('32') != -1:
                data.append('32')
            elif chinese.find('33') != -1:
                data.append('33')
            else:
                data.append('no_kaizhuang')
        for x in val:
            if x.find('×') != -1 or x.find('x') != -1 or x.find('X') != -1:
                x = re.sub(r'mm2', '', x)
                x = re.sub(r'X', '×', x)
                x = re.sub(r'x', '×', x)
                if x[-1] == '0' and x[-2] == '.':
                    x = x[:-2]
                guige = x
                break
        if guige == '':
            n = ''
            s = ''
            for x in val:
                if x.find('线芯数') != -1:
                    n = "".join(filter(lambda s: s in '0123456789', x))
                if x.find('线芯截面积') != -1:
                    s = "".join(filter(lambda s: s in '0123456789.', x))
            guige = n + '×' + s
            if guige  == '×':
                guige = 'error'
            if guige[-1] == '0' and guige[-2] == '.':
                guige = guige[:-2]
        data.append(guige)
        final.append(data)
        temp = []
        for i in final:
            t = []
            for m in i:
                x = re.sub(' ','',m)
                t.append(x)
            temp.append(t)
        final = temp
    return final,cu

def datacreat(x):
    final = []
    for data in x:
        print('data',data)
        guige = data[-1]
        print('guige',guige)
        temp = []
        if len(data) != 12:
            temp = ['error']
        else:
            if data[3] != 'none':
                xinghao = 'K' + data[0] + 'P'
            else:
                xinghao = 'K' + data[0]
            temp.append(guige)
            temp.append(xinghao)
            temp.append(data[1])
            temp.append(data[2])
            temp.append(data[3])
            temp.append(data[4])
            temp.append(data[5])
            temp.append(data[6])
            temp.append(data[7])
            temp.append(data[8])
            temp.append(data[9])
            temp.append(data[10])
        final.append(temp)
    return final



def getprice(base,data,cu,file,_):
    c = []
    line_of_price = []
    final_list = []
    if cu > 25000 and cu <= 30000:
        line_of_price = [5,6]
        c = [25000,30000]
    elif cu > 30000 and cu <= 35000:
        line_of_price = [6,7]
        c = [30000,35000]
    elif cu > 35000 and cu <= 40000:
        line_of_price = [7,8]
        c =[35000,40000]
    elif cu > 40000 and cu <= 45000:
        line_of_price = [8,9]
        c = [40000,45000]
    elif cu > 45000 and cu <= 50000:
        line_of_price = [9,10]
        c = [45000,50000]
    elif cu > 50000 and cu <= 55000:
        line_of_price = [10,11]
        c = [50000,55000]
    elif cu > 55000 and cu <= 60000:
        line_of_price = [11,12]
        c = [55000,60000]
    elif cu > 60000 and cu <= 65000:
        line_of_price = [12,13]
        c = [60000,65000]
    elif cu > 65000 and cu <= 70000:
        line_of_price = [13,14]
        c = [65000,70000]
    elif cu > 70000 and cu <= 75000:
        line_of_price = [14,15]
        c = [70000,75000]
    elif cu > 75000 and cu <= 80000:
        line_of_price = [15,16]
        c = [75000,80000]
    sheet = base.active
    # print(data)
    for d in data:
        if d == ['error']:
            final_list.append('error')
            continue
        price_list = []
        guige = [d[1],d[0]]
        # print(guige)
        h = 0
        for m in range(sheet.min_row,sheet.max_row+1):
            h += 1
            cell = sheet.cell(m,4).value
            # print(cell)
            if cell == guige[1]:
                if sheet.cell(h,2).value == guige[0]:
                    price1 = sheet.cell(h,line_of_price[0]+1).value
                    price2 = sheet.cell(h,line_of_price[1]+1).value
                    priceplus = 0.0
                    if d[5] == 'anti_fire':
                        priceplus += sheet.cell(h, 18).value
                    if d[2] == 'A':
                        priceplus += sheet.cell(h, 19).value
                    elif d[2] == 'B':
                        priceplus += sheet.cell(h, 20).value
                    elif d[2] == 'C':
                        priceplus += sheet.cell(h, 21).value
                    if d[-3] == 'no_smoke':
                        priceplus += sheet.cell(h, 22).value
                    if d[-1] == '22':
                        priceplus += sheet.cell(h, 24).value
                    elif d[-1] == '23':
                        priceplus += sheet.cell(h, 25).value
                    elif d[-1] == '32':
                        priceplus += sheet.cell(h, 26).value
                    elif d[-1] == '33':
                        priceplus += sheet.cell(h, 27).value
                    if d[6] == 'anti_ant':
                        priceplus += sheet.cell(h, 30).value
                    if d[8] == 'anti_cold':
                        priceplus += sheet.cell(h, 31).value
                    if d[4] == 'p1':
                        priceplus += 0
                    elif d[4] == 'p2':
                        priceplus += sheet.cell(h, 32).value
                    elif d[4] == 'p3':
                        priceplus += sheet.cell(h, 33).value
                    elif d[4] == 'p4':
                        priceplus += sheet.cell(h, 34).value
                    if d[7] == 'anti_mouse':
                        priceplus += sheet.cell(h, 37).value
                    if d[3] == 'soft_heart':
                        priceplus += sheet.cell(h, 23).value
                    price_list.append(h)
                    price_list.append(price1)
                    price_list.append(price2)
                    the = cu - c[0]
                    price = (price2 - price1) / 5000 * the + price1 + priceplus
                    price_list.append(price)
                    final_list.append(price_list)
                    break
        if price_list == []:
            price_list = 'error'
            final_list.append(price_list)
    return final_list

def toexcel(pricelist,file,_,path):
    sheet = file.active
    n = sheet.max_column
    for i in range(len(pricelist)):
        # print(pricelist[i])
        if pricelist[i] != 'error':
            sheet.cell(i+4, n + 1, pricelist[i][0])
            sheet.cell(i+4, n + 2, pricelist[i][1])
            sheet.cell(i+4, n + 3, pricelist[i][2])
            sheet.cell(i+4, n + 4, pricelist[i][3])
        else:
            sheet.cell(i + 4, n + 1, '出错')
            sheet.cell(i + 4, n + 2, '出错')
            sheet.cell(i + 4, n + 3, '出错')
            sheet.cell(i + 4, n + 4, '出错')
        file.save(filename=path)



def control():
    file, _, path = get_path()
    data, cu = getdata(file)
    print(data)
    print('输入基础表地址')
    data = datacreat(data)
    print(data)
    base, a, path1 = get_path()
    x = getprice(base, data, cu, file, _)
    print(x)
    toexcel(x, file, _, path)
    print('写入完成')

