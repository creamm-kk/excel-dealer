import openpyxl
import xlrd
import re
from xlutils.copy import copy

def get_path():
    path = input('输入文件路径：')
    path = path.split('\"')[1]
    if path[-3:] == 'xls':
        workbook = xlrd.open_workbook(path)
        return workbook,0,path
    else:
        workbook = openpyxl.load_workbook(path)
        return workbook,1,path


def not_empty(s):
    return s and s.strip()

def getdata0(file):
    number = input('输入你要提取的格子数量和第一个格子的编号(两个数字表示)，格式为:1,3,2 \n记得在英文模式输入，第一个数是行数-1，第二个是列数-1\n')
    sheet = file[0]
    num = number.split(',')[0]
    location = number.split(',')[1:3]
    loc = [int(location[0]),int(location[1])]
    datas = []
    dat = []
    final = []
    for i in range(int(num)):
        data = sheet.cell_value(loc[0],loc[1])
        # data = re.sub('[\u4e00-\u9fa5]', '', data)
        datas.append(data)
        loc[0] += 1
    for x in datas:
        dat = [] #型号，耐火，阻燃等级，低烟无卤，防白蚁，低温，线芯数，相线截面，中线截面，地线截面，
        if x.find('YJV') != -1:
            if x.find('BPYJVP') != -1:
                dat.append('BPYJVP')
            else:
                dat.append('YJV')
        if x.find('YJY') != -1:
            if x.find('BPYJYP') != -1:
                dat.append('BPYJYP')
            else:
                dat.append('YJY')
        if x.find('GG') != -1:
            if x.find('BPYJYP') != -1:
                dat.append('BPGGP')
            else:
                dat.append('GG')
        #以上是型号搜索
        if x.find('NH') != -1 or x.find('耐火') != -1 or x.find('N') != -1:
            dat.append('NH')
        else:
            dat.append('NNH')
        #以上是耐火
        if x.find('A') != -1:
            dat.append('A')
        elif x.find('B') != -1:
            dat.append('B')
        else:
            dat.append('C')
        #以上是阻燃
        if x.find('低烟无卤') != -1:
            dat.append('low_smoke')
        else:
            dat.append('smoke')
        #以上是低烟无卤
        if x.find('22') != -1:
            dat.append('22')
        elif x.find('23') != -1:
            dat.append('23')
        elif x.find('32') != -1:
            dat.append('32')
        elif x.find('33') != -1:
            dat.append('33')
        else:
            dat.append('no_sheild')
        #以上是铠装
        if x.find('白蚁') != -1:
            dat.append('anti_ant')
        else:
            dat.append('ant_right')
        #以上是防白蚁
        if x.find('低温') != -1:
            dat.append('anti_cold')
        else:
            dat.append('cold_right')
        #以上是耐低温
        flag = 0
        print(x)
        if x.find('×') != -1 or x.find('x') != -1:
            flag = 1 #转简单模式

        if flag == 1:
            x = re.sub(r"x", '×', x)
            x = re.sub('mm2', '', x)
            x = re.sub('[\u4e00-\u9fa5]', '', x)
            x = "".join(filter(lambda s: s in '0123456789+×.,kK/-， ', x))
            x = re.split(',|\s|K|k|/|-|，', x)
            x = filter(not_empty, x)
            x = list(x)
            if len(x[-1]) >= 8:
                x = x[:-1]
            o = 0
            while x[o].find('×'):
                dat.append(x[o])
                break
        else:
            x = re.sub('mm2', '', x)
            x = "".join(filter(lambda s: s in '0123456789×+.,kK-/， ', x))
            x = re.split(',|\s|K|k|-|/|，', x)
            x = filter(not_empty, x)
            x = list(x)
            # print('异常',x)
            i = 0
            qizi = 0
            num_of_line = ''
            temp = []
            for em in x:
                if len(em)<8:
                    temp.append(em)
            x = temp
            for row in x:
                loca = row.find('+')
                if loca == -1:
                    qizi = 0
                    i += 1
                    continue
                else:
                    qizi = 1
                    num_of_line = x[i]
                    break
            guige = ''
            # print(x)
            # print(num_of_line)
            if qizi == 1:
                if num_of_line.find('+') != -1:
                    numb = int(eval(num_of_line))
                else:
                    numb = int(num_of_line)
                if numb == 5:
                    if x[i+1] == x[i+2]:
                        guige = '5×' + x[i + 1]
                    else:
                        guige = '3×' + x[i + 1] + '+2×' + x[i + 2]
                elif numb == 4:
                    if x[i + 1] == x[i + 2]:
                        guige = '4×' + x[i + 1]
                    else:
                        guige = '3×' + x[i + 1] + '+1×' + x[i + 2]
                elif numb == 3:
                    guige = '3×' + x[i + 1]
                else:
                    print('线芯数有问题')
                dat.append(guige)
            else:
                if x[-4] == '5':
                    if x[-3] == x[-2]:
                        guige = '5×' + x[-3]
                    else:
                        guige = '3×' + x[-3] + '+2×' + x[-2]
                elif x[-4] == '3':
                    guige = '3×' + x[-3]
                elif x[-3] == '4':
                    if x[-2] == x[-1]:
                        guige = '4×' + x[-2]
                    else:
                        guige = '3×' + x[-2] + '+1×' + x[-1]
                elif x[-2] == '3':
                    guige = '3×'+ x[-1]
                dat.append(guige)
        final.append(dat)
    cu = sheet.cell(1,0).value
    return final,cu

def getdata1(file):
    # sheet_name = input('输入所在sheet的名字：')
    number = input('输入你要提取的格子数量和第一个格子的编号，格式为:1,C4 \n记得要大写,并在英文模式输入\n')
    # sheet = file[sheet_name]
    sheet = file.active
    num = number.split(',')[0]
    location = number.split(',')[1]
    datas = []
    final = []
    for i in range(int(num)):
        data = sheet[location].value
        datas.append(data)
        raw = re.sub(r"[A-Za-z]","",location)
        raw = int(raw)
        raw += 1
        location = re.sub(r"[\d]","",location)+str(raw)
    for x in datas:
        dat = []  # 型号，耐火，阻燃等级，低烟无卤，防白蚁，低温，线芯数，相线截面，中线截面，地线截面，
        if x.find('YJV') != -1:
            if x.find('BPYJVP') != -1:
                dat.append('BPYJVP')
            else:
                dat.append('YJV')
        if x.find('YJY') != -1:
            if x.find('BPYJYP') != -1:
                dat.append('BPYJYP')
            else:
                dat.append('YJY')
        if x.find('GG') != -1:
            if x.find('BPYJYP') != -1:
                dat.append('BPGGP')
            else:
                dat.append('GG')
        # 以上是型号搜索
        if x.find('NH') != -1 or x.find('耐火') != -1 or x.find('N') != -1:
            dat.append('NH')
        else:
            dat.append('NNH')
        # 以上是耐火
        if x.find('A') != -1:
            dat.append('A')
        elif x.find('B') != -1:
            dat.append('B')
        else:
            dat.append('C')
        # 以上是阻燃
        if x.find('低烟无卤') != -1:
            dat.append('low_smoke')
        else:
            dat.append('smoke')
        # 以上是低烟无卤
        if x.find('22') != -1 and x.find('mm22') == -1:
            dat.append('22')
        elif x.find('23') != -1 and x.find('mm23') == -1:
            dat.append('23')
        elif x.find('32') != -1:
            dat.append('32')
        elif x.find('33') != -1:
            dat.append('33')
        else:
            dat.append('no_sheild')
        # 以上是铠装
        if x.find('白蚁') != -1:
            dat.append('anti_ant')
        else:
            dat.append('ant_right')
        # 以上是防白蚁
        if x.find('低温') != -1:
            dat.append('anti_cold')
        else:
            dat.append('cold_right')
        # 以上是耐低温
        flag = 0
        print(x)
        if x.find('×') != -1 or x.find('x') != -1 or x.find('X') != -1:
            flag = 1  # 转简单模式
        print(flag)
        if flag == 1:
            x = re.sub(r"x", '×', x)
            x = re.sub(r"X", '×', x)
            x = re.sub('0.6/1KV', '', x)
            x = re.sub('0.6/1kV', '', x)
            x = re.sub('0.6/1Kv', '', x)
            x = re.sub('0.6/1kv', '', x)
            x = re.sub('YJV22', '', x)
            x = re.sub('YJV32', '', x)
            x = re.sub('YJV23', '', x)
            x = re.sub('YJV33', '', x)
            x = re.sub('YJY22', '', x)
            x = re.sub('YJY32', '', x)
            x = re.sub('YJY23', '', x)
            x = re.sub('YJY33', '', x)
            x = re.sub('GG22', '', x)
            x = re.sub('GG32', '', x)
            x = re.sub('GG23', '', x)
            x = re.sub('GG33', '', x)
            x = re.sub('[\u4e00-\u9fa5]', '', x)
            x = "".join(filter(lambda s: s in '0123456789×+.,mkK-/， ', x))
            x = re.split(',|\s|K|k|-|/|，|mm2', x)
            x = filter(not_empty, x)
            x = list(x)
            if len(x[-1]) >= 8 and x[-1].find('×') == -1:
                x = x[:-1]
            for a in x:
                if a.find('×') != -1:
                    dat.append(a)
                    print('after',a)
                    break
        else:
            x = re.sub('mm2', '', x)
            x = re.sub('0.6/1KV','',x)
            x = re.sub('0.6/1kV','',x)
            x = re.sub('0.6/1Kv','',x)
            x = re.sub('0.6/1kv','',x)
            x = re.sub('YJV22','',x)
            x = re.sub('YJV32','',x)
            x = re.sub('YJV23','',x)
            x = re.sub('YJV33','',x)
            x = re.sub('YJY22', '', x)
            x = re.sub('YJY32', '', x)
            x = re.sub('YJY23', '', x)
            x = re.sub('YJY33', '', x)
            x = re.sub('GG22', '', x)
            x = re.sub('GG32', '', x)
            x = re.sub('GG23', '', x)
            x = re.sub('GG33', '', x)
            x = "".join(filter(lambda s: s in '0123456789×+.,kK-/， ', x))
            x = re.split(',|\s|K|k|-|/|，', x)
            x = filter(not_empty, x)
            x = list(x)
            print('after',x)
            i = 0
            qizi = 0
            num_of_line = ''
            temp = []
            tem = 0
            for em in x:
                if len(em) < 8:
                    temp.append(em)
            x = temp
            for row in x:
                loca = row.find('+')
                if loca == -1:
                    qizi = 0
                    i += 1
                    continue
                elif loca == 1:
                    qizi = 1
                    num_of_line = x[i]
                    break
                else:
                    qizi = 2
                    tem = i
                    break
            guige = ''
            # print(num_of_line)
            print(qizi)
            if qizi == 1:
                if num_of_line.find('+') != -1:
                    numb = int(eval(num_of_line))
                else:
                    numb = int(num_of_line)
                if numb == 5:
                    if x[i + 1] == x[i + 2]:
                        guige = '5×' + x[i + 1]
                    else:
                        guige = '3×' + x[i + 1] + '+2×' + x[i + 2]
                elif numb == 4:
                    if x[i + 1] == x[i + 2]:
                        guige = '4×' + x[i + 1]
                    else:
                        guige = '3×' + x[i + 1] + '+1×' + x[i + 2]
                elif numb == 3:
                    guige = '3×' + x[i + 1]
                else:
                    print('线芯数有问题')
                dat.append(guige)
            elif qizi == 0:
                if len(x)>=4:
                    if x[-4] == '5':
                        if x[-3] == x[-2]:
                            guige = '5×' + x[-3]
                        else:
                            guige = '3×' + x[-3] + '+2×' + x[-2]
                    elif x[-4] == '3':
                        guige = '3×' + x[-3]
                    elif x[-4] == '4':
                        guige = '4×' + x[-3]
                    elif x[-3] == '4':
                        if x[-2] == x[-1]:
                            guige = '4×' + x[-2]
                        else:
                            guige = '3×' + x[-2] + '+1×' + x[-1]
                    elif x[-2] == '3':
                        guige = '3×' + x[-1]
                    dat.append(guige)
                elif len(x)>=3:
                    if x[-3] == '4':
                        if x[-2] == x[-1]:
                            guige = '4×' + x[-2]
                        else:
                            guige = '3×' + x[-2] + '+1×' + x[-1]
                    elif x[-2] == '3':
                        guige = '3×' + x[-1]
                    dat.append(guige)
                else:
                    if x[-2] == x[-1]:
                        guige = '4×' + x[-2]
                    else:
                        guige = '3×' + x[-2] + '+1×' + x[-1]
                    if x[-2] == '3':
                        guige = '3×' + x[-1]
                    dat.append(guige)
            else:
                print(x[tem])
                dat.append(x[tem])

        final.append(dat)
    cu = sheet.cell(2, 1).value
    return final, cu


def getprice(base,data,cu,file,_):
    cu = int(cu)
    cu = cu // 1000
    cu *= 1000
    c = []
    line_of_price = []
    final_list = []
    if cu > 25000 and cu <= 30000:
        line_of_price = [5,6]
        c = [25000,30000]
    elif cu > 30000 and cu <= 35000:
        line_of_price = [6,7]
        c = [30000,35000]
    elif cu > 35000 and cu <= 40000:
        line_of_price = [7,8]
        c =[35000,40000]
    elif cu > 40000 and cu <= 45000:
        line_of_price = [8,9]
        c = [40000,45000]
    elif cu > 45000 and cu <= 50000:
        line_of_price = [9,10]
        c = [45000,50000]
    elif cu > 50000 and cu <= 55000:
        line_of_price = [10,11]
        c = [50000,55000]
    elif cu > 55000 and cu <= 60000:
        line_of_price = [11,12]
        c = [55000,60000]
    elif cu > 60000 and cu <= 65000:
        line_of_price = [12,13]
        c = [60000,65000]
    elif cu > 65000 and cu <= 70000:
        line_of_price = [13,14]
        c = [65000,70000]
    elif cu > 70000 and cu <= 75000:
        line_of_price = [14,15]
        c = [70000,75000]
    elif cu > 75000 and cu <= 80000:
        line_of_price = [15,16]
        c = [75000,80000]
    sheet = base.active
    for d in data:
        price_list = []
        guige = [d[0],d[-1]]
        h = 0
        for m in range(sheet.min_row,sheet.max_row+1):
            h += 1
            cell = sheet.cell(m,4).value
            if cell == guige[1]:
                if sheet.cell(h,2).value == guige[0]:
                    price1 = sheet.cell(h,line_of_price[0]+1).value
                    price2 = sheet.cell(h,line_of_price[1]+1).value
                    priceplus = 0.0
                    if d[1] == 'NH':
                        priceplus += sheet.cell(h, 18).value
                    if d[2] == 'A':
                        priceplus += sheet.cell(h, 19).value
                    elif d[2] == 'B':
                        priceplus += sheet.cell(h, 20).value
                    elif d[2] == 'C':
                        priceplus += sheet.cell(h, 21).value
                    if d[3] == 'low_smoke':
                        priceplus += sheet.cell(h, 22).value
                    if d[4] == '22':
                        priceplus += sheet.cell(h, 23).value
                    elif d[4] == '23':
                        priceplus += sheet.cell(h, 24).value
                    elif d[4] == '32':
                        priceplus += sheet.cell(h, 25).value
                    elif d[4] == '33':
                        priceplus += sheet.cell(h, 26).value
                    if d[5] == 'anti_ant':
                        priceplus += sheet.cell(h, 33).value
                    if d[6] == 'anti_cold':
                        priceplus += sheet.cell(h, 34).value
                    price_list.append(h)
                    price_list.append(price1)
                    price_list.append(price2)
                    the = cu-c[0]
                    price = (price2-price1)/5000 * the + price1 + priceplus
                    price_list.append(price)
                    final_list.append(price_list)
                    break
        if price_list == []:
            price_list = 'error'
            final_list.append(price_list)
    return final_list

def toexcel(pricelist,file,_,path):
    sheet = file.active
    n = sheet.max_column
    for i in range(len(pricelist)):
        # print(pricelist[i])
        if pricelist[i] != 'error':
            sheet.cell(i+4, n + 1, pricelist[i][0])
            sheet.cell(i+4, n + 2, pricelist[i][1])
            sheet.cell(i+4, n + 3, pricelist[i][2])
            sheet.cell(i+4, n + 4, pricelist[i][3])
        else:
            sheet.cell(i + 4, n + 1, '出错')
            sheet.cell(i + 4, n + 2, '出错')
            sheet.cell(i + 4, n + 3, '出错')
            sheet.cell(i + 4, n + 4, '出错')
        file.save(filename=path)

def electric():
    file,_,path= get_path()
    if _ == 0:
        print('你正在使用xls旧版，因难以写入放弃优化，建议改为xlsx新版')
        x = input('若要继续使用请按0和回车')
        while x == "0":
            break
        data,cu = getdata0(file)
    else:
        data,cu = getdata1(file)
    print('输入基础表地址')
    print(data)
    base,a,path1 = get_path()
    cu = cu[-5:]
    x = getprice(base,data,cu,file,_)
    print(x)
    toexcel(x,file,_,path)
    print('写入完成')





