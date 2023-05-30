import openpyxl
import xlrd
import re

def qiaojia(val,sheet):
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    for m in range(sheet.min_row,sheet.max_row+1):
        h += 1
        cell1 = sheet.cell(m,3).value
        cell2 = sheet.cell(m,2).value
        if cell1 == whh and cell2 == val[-1]:
            if val[2] == 'tishi':
                price += float(sheet.cell(h,4).value)
            elif val[2] == 'caoshi':
                price += float(sheet.cell(h,5).value)
            else:
                return ['错误','错误']
            if val[3] == 'yes':
                price += float(sheet.cell(h,6).value)
            if val[4] == '2tibian':
                price += float(sheet.cell(h,7).value)
            break
    return price,h
def baoku(val, sheet):
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 3).value
        cell2 = sheet.cell(m, 2).value
        if cell1 == whh and cell2 == val[-1]:
            price += sheet.cell(m, 9).value
            break
    return price, h
def fengtou(val, sheet):
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 3).value
        cell2 = sheet.cell(m, 2).value
        if cell1 == whh and cell2 == val[-1]:
            price += sheet.cell(m, 8).value
            break
    return price, h
def gaiban(val, sheet):
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 3).value
        cell2 = sheet.cell(m, 2).value
        if cell1 == whh and cell2 == val[-1]:
            price += sheet.cell(m, 6).value
            break
    return price, h
def guanjietou(val, sheet):
    h = 0
    price = 0
    if val[3] == 'buxiugang':
        name = '铝合金桥架管接头'
    else:
        name = '不锈钢桥架管接头'
    xinghao = val[0]
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 3).value
        cell2 = sheet.cell(m, 2).value
        if cell1 == xinghao and cell2 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def zadai(val, sheet):
    h = 0
    price = 0
    name = '扎带'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 3).value
        cell2 = sheet.cell(m, 2).value
        if cell1 == val[0] + '×0.5' and cell2 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def geban(val, sheet):
    h = 0
    price = 0
    name = '铝合金桥架隔板'
    height = val[0]
    houdu = val[1]
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        try:
            cell2 = re.split(',|mm| |/|-|×|，', cell2)
        except:
            continue
        num_list = []
        for data in cell2:
            try:
                n = "".join(filter(lambda s: s in '0123456789', data))
                if n != '':
                    num_list.append(n)
            except:
                continue
        if len(num_list) < 2:
            continue
        height1 = num_list[0]
        houdu1 = num_list[1]
        if height1 == height and houdu1 == houdu and cell1 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def zhijiepian(val, sheet):
    h = 0
    price = 0
    name = '铝合金直接片'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        try:
            cell2 = "".join(filter(lambda s: s in '0123456789', cell2))
        except:
            continue
        if val[0] == cell2  and cell1 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def wanjiepian(val, sheet):
    h = 0
    price = 0
    name = '铝合金弯接片'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        try:
            cell2 = "".join(filter(lambda s: s in '0123456789', cell2))
        except:
            continue
        if val[0] == cell2  and cell1 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def tiaojiaopian(val, sheet):
    h = 0
    price = 0
    name = '铝合金调角片'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        try:
            cell2 = "".join(filter(lambda s: s in '0123456789', cell2))
        except:
            continue
        if val[0] == cell2  and cell1 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def lianjiexian(val, sheet):
    h = 0
    price = 0
    name = '桥架用跨接连线'
    height = val[0]
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        try:
            cell2 = re.split(',|mm|mm2| |/|-|×|1×|1x|1X|，', cell2)
        except:
            continue
        num_list = []
        for data in cell2:
            try:
                n = "".join(filter(lambda s: s in '0123456789', data))
                if n != '':
                    num_list.append(n)
            except:
                continue
        if len(num_list) < 1:
            continue
        height1 = num_list[0]
        if height1 == height and cell1 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h
def gudingyaban(val, sheet):
    h = 0
    price = 0
    price_list = ''
    name = '铝合金调角片'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        if cell1 == name:
            price += sheet.cell(m, 4).value
            price_list += str(price)
            price_list += ','
    return price_list, h
def xiangjiaodian(val, sheet):
    h = 0
    price = 0
    price_list = ''
    name = '桥架用防电化学橡胶垫'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 2).value
        cell2 = sheet.cell(m, 3).value
        if cell1 == name:
            price += sheet.cell(m, 4).value
            break
    return price, h