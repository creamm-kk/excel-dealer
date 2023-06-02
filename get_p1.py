import openpyxl
import xlrd
import re

def qiaojia(val,sheet,rank): #rank从1开始
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    # print(sheet.min_row,sheet.max_row)
    for m in range(sheet.min_row,sheet.max_row+1):
        flag = 1
        h += 1
        cell1 = sheet.cell(m,5).value
        # print(cell1)
        if cell1 != None and cell1.find(whh) != -1 and cell1.find(val[-1]) != -1 and cell1.find(val[-3])!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            if val[2] == 'tishi':
                if cell1.find('梯式') != -1:
                    flag = 1
                else:
                    continue
            elif val[2] == 'caoshi':
                if cell1.find('槽式') != -1:
                    flag = 1
                else:
                    continue
            else:
                return ['错误','错误']
            if val[3] == 'yes':
                if cell1.find('带盖板') != -1:
                    flag = 1
                else:
                    continue
            elif val[3] == 'no':
                if cell1.find('无盖板') != -1:
                    flag = 1
                else:
                    continue
            else:
                return ['错误', '错误']
            if val[4] == '2tibian':
                if cell1.find('双梯边') != -1:
                    flag = 1
                else:
                    continue
            elif val[4] == '1tibian':
                if cell1.find('双梯边') == -1:
                    flag = 1
                else:
                    continue
            price += float(sheet.cell(m,rank+6).value)
            print(price,h)
            return price,h
    return ['错误', '错误']
    #     cell2 = sheet.cell(m,2).value
    #     if cell1 == whh and cell2 == val[-1]:
    #         if val[2] == 'tishi':
    #             price += float(sheet.cell(h,4).value)
    #         elif val[2] == 'caoshi':
    #             price += float(sheet.cell(h,5).value)
    #         else:
    #             return ['错误','错误']
    #         if val[3] == 'yes':
    #             price += float(sheet.cell(h,6).value)
    #         if val[4] == '2tibian':
    #             price += float(sheet.cell(h,7).value)
    #         break
    # return price,h
def baoku(val, sheet, rank):
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        if cell1 != None and cell1.find(whh) != -1 and cell1.find('抱箍')!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
        # cell2 = sheet.cell(m, 2).value
        # if cell1 == whh and cell2 == val[-1]:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def fengtou(val, sheet, rank):
    h = 0
    price = 0
    bihou = val[1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    print(whh)
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        # cell2 = sheet.cell(m, 2).value
        if cell1 != None and cell1.find(whh)!=-1 and (cell1.find('封头')!=-1 or cell1.find('终端头')!=-1) and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def gaiban(val, sheet, rank):
    h = 0
    price = 0
    bihou = val[1]
    if val[-1]=='直通桥架':
        name = '直通'
    elif val[-1]=='水平三通':
        name = '三通'
    elif val[-1] == '水平四通':
        name = '四通'
    else:
        name = val[-1]
    if bihou % 1 == 0:
        bihou = str(bihou)
        bihou = bihou[:-2]
    else:
        bihou = str(bihou)
    whh = val[0] + '×' + bihou
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    print(whh)
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        # print(cell1)
        # cell2 = sheet.cell(m, 2).value
        if cell1 != None and cell1.find(whh)!=-1 and cell1.find('盖板')!=-1 and cell1.find('带盖板')==-1 and cell1.find('无盖板')==-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            print(sheet.cell(m, rank+6).value)
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def guanjietou(val, sheet, rank):
    h = 0
    price = 0
    if val[3] == 'buxiugang':
        name = '铝合金桥架管接头'
    else:
        name = '不锈钢桥架管接头'
    xinghao = val[0]
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        # cell2 = sheet.cell(m, 2).value
        if cell1 != None and cell1.find(xinghao)!=-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, 6+rank).value
            break
    return price, h
def zadai(val, sheet, rank):
    h = 0
    price = 0
    name = '扎带'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        if val[-3] == '':
            fpen = '喷'
        else:
            fpen = '???'
        # cell2 = sheet.cell(m, 2).value
        if cell1 != None and cell1.find(val[0] + '×0.5')!=-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def geban(val, sheet, rank):
    h = 0
    price = 0
    name = '桥架隔板'
    height = val[0]
    houdu = val[1]
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        # cell2 = sheet.cell(m, 3).value
        # try:
        #     cell2 = re.split(',|mm| |/|-|×|，', cell2)
        # except:
        #     continue
        # num_list = []
        # for data in cell2:
        #     try:
        #         n = "".join(filter(lambda s: s in '0123456789', data))
        #         if n != '':
        #             num_list.append(n)
        #     except:
        #         continue
        # if len(num_list) < 2:
        #     continue
        # height1 = num_list[0]
        # houdu1 = num_list[1]
        if cell1 != None and cell1.find(height+'×'+houdu)!=-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def zhijiepian(val, sheet, rank):
    h = 0
    price = 0
    name = '直接片'
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        if cell1 != None and cell1.find(val[0])!=-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def wanjiepian(val, sheet, rank):
    h = 0
    price = 0
    name = '弯接片'
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        if cell1 != None and cell1.find(val[0])!=-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def tiaojiaopian(val, sheet, rank):
    h = 0
    price = 0
    name = '调角片'
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        if cell1 != None and cell1.find(val[0])!=-1 and cell1.find(name)!=-1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank+6).value
            break
    return price, h
def lianjiexian(val, sheet, rank):
    h = 0
    price = 0
    name = '桥架用跨接连线'
    name1 = '跨连线'
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    # height = val[0]
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        if cell1 != None and cell1.find(val[0]) != -1 and (cell1.find(name) != -1 or cell1.find(name1) != -1) and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, rank + 6).value
            break
    return price, h
def gudingyaban(val, sheet, rank):
    h = 0
    price = 0
    price_list = ''
    name = '固定压板'
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        # cell2 = sheet.cell(m, 3).value
        if cell1 != None and cell1.find(name) != -1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, 6+rank).value
            price_list += str(price)
            price_list += ','
    return price_list, h
def xiangjiaodian(val, sheet, rank):
    h = 0
    price = 0
    price_list = ''
    name = '桥架用防电化学橡胶垫'
    if val[-3] == '':
        fpen = '喷'
    else:
        fpen = '???'
    for m in range(sheet.min_row, sheet.max_row + 1):
        h += 1
        cell1 = sheet.cell(m, 5).value
        # cell2 = sheet.cell(m, 3).value
        if cell1 != None and cell1.find(name) != -1 and cell1.find(val[-3])!=-1 and cell1.find(fpen)==-1:
            price += sheet.cell(m, 6+rank).value
            break
    return price, h