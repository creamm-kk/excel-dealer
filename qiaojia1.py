import openpyxl
import xlrd
import re
from xlutils.copy import copy
import fujian
import get_p1


def get_path():
    path = input('输入文件路径：')
    path = path.split('\"')[1]
    if path[-3:] == 'xls':
        workbook = xlrd.open_workbook(path)
        return workbook,0,path
    else:
        workbook = openpyxl.load_workbook(path)
        return workbook,1,path


def not_empty(s):
    return s and s.strip()

def is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return 1

    return 0

def getdata(file):
    number = input('输入你要提取的格子数量和第一个格子的编号，格式为:1,C4 \n记得要大写,并在英文模式输入\n')
    sheet = file.active
    cu = sheet.cell(2, 1).value[-5:]
    cu = int(cu)
    num = number.split(',')[0]
    location = number.split(',')[1]
    datas = []
    raw = re.sub(r"[A-Za-z]", "", location)
    raw = int(raw)
    for i in range(int(num)):#读取本体
        location = re.sub(r"[\d]","",location)+str(raw)
        location1 = 'J' + str(raw)#梯边
        location2 = 'K' + str(raw)#壁厚
        tibian = str(sheet[location1].value)
        bihou = str(sheet[location2].value)
        print(location,location1,location2)
        print(type(bihou))
        data = sheet[location].value + ',' + tibian + ',' + bihou
        a = re.sub(u"\\（.*?）|\\{.*?}|\\[.*?]|\\【.*?】", "", data)
        data = re.sub(u"\\(.*?\\)|\\{.*?}|\\[.*?]", "", a)
        datas.append(data)
        raw += 1

    return datas,cu

def data_dealer(datas):
    features = ''
    features_list = []
    for val in datas:
        #先处理桥架本体
        if val.find('异径') != -1:
            classes = 'error'
            features = ['', '', '', '', '', 'error', '']
            features_list.append(features)
        elif val.find('直通') != -1 or val.find('弯通') != -1 or val.find('三通') != -1 or val.find('四通') != -1 or val.find('凹') != -1 or val.find('异径接头') != -1 or val.find('凸') != -1 and val.find('抱箍') == -1 and val.find('封头') == -1 and val.find('管接头') == -1 and val.find('扎带') == -1 and val.find('隔板') == -1 and val.find('直接片') == -1 and val.find('连接板') == -1 and val.find('弯接片') == -1 and val.find('调角片') == -1 and val.find('调宽片') == -1 and val.find('调高片') == -1 and val.find('连接线') == -1 and val.find('接地线') == -1 and val.find('跨接线') == -1 and val.find('接连线') == -1 and val.find('固定压板') == -1 and val.find('胶垫') == -1:
            val = re.sub(r'X', '×', val)
            val = re.sub(r'x', '×', val)
            val = re.sub(r'\*', '×', val)
            val = re.split(',|mm|mm2| |/|-|×|，', val)
            val = filter(not_empty, val)
            val = list(val)
            print(val)
            wh = get_wh(val)
            wh = wh[0] + '×' + wh[1]
            # print(wh)
            features = itself(val,wh)
            print(features)
            features_list.append(features)
        else:
            print('是配件')
            val = re.sub(r'X', '×', val)
            val = re.sub(r'x', '×', val)
            val = re.sub(r'\*', '×', val)
            vals = re.split(',|mm|mm2| |/|-|×|，', val)
            vals = filter(not_empty, vals)
            vals = list(vals)
            classes = ''
            if val.find('异径') != -1:
                classes = 'error'
                features = ['', '', '', '', '', 'error','']
            elif val.find('抱箍') != -1:
                classes = 'baoku'
                features = fujian.baoku(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('封头') != -1 or val.find('终端头') != -1:
                classes = 'fengtou'
                features = fujian.fengtou(vals)
                # print('aaaa',features)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                # print('aaaa',features)
                features.insert(-2, pen)
                # print(features)
            elif val.find('管接头') != -1:
                classes = 'guanjietou'
                features = fujian.guanjietou(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('扎带') != -1:
                classes = 'zadai'
                features = fujian.zadai(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('隔板') != -1:
                classes = 'geban'
                features = fujian.geban(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('直接片') != -1 or val.find('伸缩片') != -1 or val.find('连接片') != -1 or val.find('连接板') != -1:
                classes = 'zhijiepian'
                features = fujian.zhijiepian(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('弯接片') != -1:
                classes = 'wanjiepian'
                features = fujian.wanjiepian(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('调角片') != -1 or val.find('调高片') != -1 or val.find('调宽片') != -1:
                classes = 'tiaojiaopian'
                features = fujian.tiaojiaopian(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('连接线') != -1 or val.find('接地线') != -1 or val.find('跨接线') != -1 or val.find('接连线') != -1:
                classes = 'lianjiexian'
                features = fujian.lianjiexian(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('固定压板') != -1:
                classes = 'gudingyaban'
                features = fujian.gudingyaban(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            elif val.find('胶垫') != -1:
                classes = 'xiangjiaodian'
                features = fujian.xiangjiaodian(vals)
                if val.find('喷') != -1 or val.find('环氧树脂') != -1:
                    pen = '喷'
                else:
                    pen = ''
                features.insert(-2, pen)
            else:
                classes = 'error'
                features = ['', '', '', '', '', 'error','']
            print(vals)
            print(features)
            features_list.append(features)
    return features_list


def itself(val,wh):
    classes = ''
    mode = ''
    have_gaiban = 'no'
    tibian = ''
    bihou = 0
    zhonglei = 'qiaojia'
    pen = ''
    for data in val[:-2]:
        if data.find('梯') != -1:
            mode = 'tishi'
        if data.find('托盘') != -1 or data.find('槽') != -1:
            mode = 'caoshi'
        if data.find('直通') != -1:
            classes = '直通桥架'
        elif data.find('三通') != -1:
            classes = '水平三通'
        elif data.find('四通') != -1:
            classes = '水平四通'
        elif data.find('水平') != -1:
            classes = '水平弯通'
        elif data.find('垂直') != -1 or data.find('凹') != -1 or data.find('凸') != -1 or data.find('异径接头') != -1:
            classes = '垂直弯通'
        if data.find('盖板') != -1 or data.find('护罩') != -1:
            if data.find('配') != -1 or data.find('有') != -1 or data.find('带') != -1:
                have_gaiban = 'yes'
            elif data.find('非盖板') != -1 or data.find('无盖板') != -1 or data.find('非护罩') != -1 or data.find('无护罩') != -1:
                have_gaiban = 'no'
            else:
                zhonglei = 'gaiban'
        if data.find('喷') != -1 or data.find('环氧树脂') != -1:
            pen = '喷'


    if val[-2] =='双' or val[-2] =='双梯边':
        tibian = '2tibian'
    else:
        tibian = '1tibian'
    if mode == '':
        mode = 'caoshi'
    bihou = float(val[-1])
    final = [wh,bihou,mode,have_gaiban,tibian,pen,zhonglei,classes]
    return final



def get_wh(val):#得到长宽
    wh = []
    for data in val:
        n = "".join(filter(lambda s: s in '0123456789', data))
        if len(n)==3 or n == '50' or n == '1000' or n == '1200':
            wh.append(n)
    wh = filter(not_empty, wh)
    wh = list(wh)
    try:
        wh = [wh[0],wh[1]]
    except:
        wh = ['0','0']
    return wh

def getprice(base, data, file, cu, _):
    sheet = base.active
    price_list = []
    i = 0
    for val in data:
        i += 1
        print(i)
        if val[-2] == 'qiaojia':
            price,h = get_p1.qiaojia(val,sheet,cu)
            if price == 0:
                prices = ['错误','错误']
            else:
                prices = [price,h]
            price_list.append(prices)
        elif val[-2] == 'baoku':
            price,h = get_p1.baoku(val,sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'fengtou':
            price, h = get_p1.fengtou(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'gaiban':
            price, h = get_p1.gaiban(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'guanjietou':
            price, h = get_p1.guanjietou(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'zadai':
            price, h = get_p1.zadai(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'geban':
            price, h = get_p1.geban(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'zhijiepian':
            price, h = get_p1.zhijiepian(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'wanjiepian':
            price, h = get_p1.wanjiepian(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'tiaojiaopian':
            price, h = get_p1.tiaojiaopian(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'lianjiexian':
            price, h = get_p1.lianjiexian(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'gudingyaban':
            price, h = get_p1.gudingyaban(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        elif val[-2] == 'xiangjiaodian':
            price, h = get_p1.xiangjiaodian(val, sheet,cu)
            if price == 0:
                prices = ['错误', '错误']
            else:
                prices = [price, h]
            price_list.append(prices)
        else:
            price_list.append(['错误','错误'])
    return price_list

def toexcel(pricelist,file,_,path):
    sheet = file.active
    n = sheet.max_column
    for i in range(len(pricelist)):
        # print(pricelist[i])
        if pricelist[i] != 'error':
            sheet.cell(i+4, n + 1, pricelist[i][0])
            sheet.cell(i+4, n + 2, pricelist[i][1])
        else:
            sheet.cell(i + 4, n + 1, '出错')
            sheet.cell(i + 4, n + 2, '出错')
        file.save(filename=path)

def qiaojia():
    file, _, path = get_path()
    data,cu = getdata(file)
    cu = 1 + (cu-17000)//500
    print(data)
    data = data_dealer(data)
    print(data)
    print('输入基础表地址')
    base, a, path1 = get_path()
    x = getprice(base, data, file, cu, _)
    print(x)
    toexcel(x, file, _, path)
    print('写入完成')


