import openpyxl
import re


def get_path():
    path = input('输入文件路径：')
    path = path.split('\"')[1]
    workbook = openpyxl.load_workbook(path)
    return workbook,1,path

def not_empty(s):
    return s and s.strip()

def is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return 1

    return 0

def getdata(file):
    number = input('输入你要提取的格子数量和第一个格子的编号，格式为:1,C3 \n记得要大写,并在英文模式输入\n')
    sheet = file.active
    num = number.split(',')[0]
    location = number.split(',')[1]
    datas = []
    final = []
    for i in range(int(num)):
        data = sheet[location].value
        datas.append(data)
        raw = re.sub(r"[A-Za-z]","",location)
        raw = int(raw)
        raw += 1
        location = re.sub(r"[\d]","",location)+str(raw)
    for val in datas:
        tot = val
        val = re.split(',|mm2| |/|-',val)
        val = filter(not_empty, val)
        val = list(val)
        print(val)
        name = ''
        W = ''
        fangshi = ''
        fangshi1 = ''
        flag = 0
        # ，，，，，，，米数（路灯），[壁式、中光强，壁式、低光强（航空障碍灯），座式、中光强，座式、低光强]x.find('支架') != -1 or
        # 灯，投光灯，泛光灯，标志灯，航空障碍灯，荧光灯，防腐荧光灯，路灯，应急灯
        ff = 0
        for x in val:
            if x.find('LED') != -1 and ff == 0:
                if x.find('投光灯') != -1:
                    name = '投光灯'
                    ff = 1
                elif x.find('泛光灯') != -1:
                    name = "节能泛光灯"
                    ff = 1
                elif x.find('标志灯') != -1:
                    name = '标志灯'
                    ff = 1
                elif x.find('航空障碍灯') != -1:
                    name = '航空障碍灯'
                    ff = 1
                elif x.find('防腐荧光灯') != -1 or tot.find('塑料') != -1:
                    name = '防腐荧光灯'
                    ff = 1
                elif x.find('荧光灯') != -1:
                    name = '荧光灯'
                    ff = 1
                elif x.find('路灯') != -1:
                    name = '路灯'
                    ff = 1
                elif x.find('应急灯') != -1:
                    name = '应急灯'
                    ff = 1
                else:
                    name = '灯'
                    ff = 1

            if x.find('W') != -1 or x.find('w') != -1 or x.find('瓦') != -1:
                x = re.sub('220V','',x)
                if name != '荧光灯':
                    loc = x.find('W') + x.find('w') + x.find('瓦') + 2
                    if loc >= 2:
                        if x[loc-1] >= '0' and x[loc-1] <= '9' and x[loc-2] >= '0' and x[loc-2] <= '9':
                            val = re.split('220V', x)
                            val = filter(not_empty, val)
                            val = list(val)
                            for v in val:
                                x = re.sub(r"[A-Za-z]", "", v)
                                if x[0] == '1' and x[1] == '×':
                                    x = x[2:]
                                if x != '':
                                    W = x + 'W'
                                    W = "".join(filter(lambda s: s in '0123456789×W', W))
                                    break
                else:
                    x = re.sub(r"x", '×', x)
                    x = re.sub(r"X", '×', x)
                    loc = x.find('W') + x.find('w') + x.find('瓦') + 2
                    if loc >= 2:
                        if x[loc-1] >= '0' and x[loc-1] <= '9' and x[loc-2] >= '0' and x[loc-2] <= '9':
                            loc2 = x.find('×')
                            mult = x[loc2-1]
                            val = re.split('220V|×', x)
                            val = filter(not_empty, val)
                            val = list(val)
                            L = 0
                            for v in val:
                                if L == 0:
                                    L += 1
                                    continue
                                x = re.sub(r"[A-Za-z]", "", v)
                                if x != '':
                                    W = mult + '×' + x + 'W'
                                    W = "".join(filter(lambda s: s in '0123456789×W', W))
                                    break

            if flag != 1 and [x.find('支架') != -1 or x.find('护栏') != -1 or x.find('吸顶') != -1 or x.find('壁') != -1 or x.find('法兰') != -1 or x.find('吊杆') != -1 or x.find('弯杆') != -1 or [[x.find('米') != -1 or x.find('m') != -1 or x.find('M') != -1] and name == '路灯']]:
                fangshi = x
                if name.find("航空") != -1:
                    if fangshi.find('壁') != -1:
                        fangshi1 = '壁式'
                        flag = 1
                    elif fangshi.find('座') != -1:
                        fangshi1 = '座式'
                        flag = 1
                    else:
                        final.append(['none','none','none'])
                        continue
                    if x.find('中光强'):
                        fangshi1 = fangshi + '中光强'
                    if x.find('低光强'):
                        fangshi1 = fangshi + '低光强'
                if fangshi.find('支架') != -1:
                    fangshi1 = '支架式'
                    flag = 1
                elif fangshi.find('护栏') != -1:
                    fangshi1 = '护栏式'
                    flag = 1
                elif fangshi.find('吸顶') != -1:
                    fangshi1 = '吸顶式'
                    flag = 1
                elif fangshi.find('壁') != -1:
                    fangshi1 = '壁式'
                    flag = 1
                elif fangshi.find('法兰') != -1:
                    fangshi1 = '法兰式'
                    flag = 1
                elif fangshi.find('吊杆') != -1:
                    fangshi1 = '吊杆式'
                    flag = 1
                elif fangshi.find('弯杆') != -1:
                    fangshi1 = '吊杆式'
                    flag = 1
                elif [[fangshi.find('米') != -1] or [fangshi.find('m') != -1] or [fangshi.find('M') != -1]] and name == '路灯':
                    loc = x.find('米') + x.find('m') + x.find('M') + 2
                    if loc >= 1:
                        m = x[loc-1]
                        if m == '0':
                            m = '10'
                        flag = 1
                        fangshi1 = '灯杆' + m + '米'
        this_one = [name,W,fangshi1]
        final.append(this_one)
    return final

def deal_cell(cell):
    cell = re.sub(r'（|）|/(|/)|LED|防水防尘防腐|高效|节能|全塑|双管|单管|三防|高度|≥|≤| |', '', cell)
    return cell

def getprice(base,data,file,_):
    final_list = []
    sheet = base.active
    for d in data:
        if d == ['error']:
            final_list.append('error')
            continue
        price_list = []
        h = 0
        for m in range(sheet.min_row,sheet.max_row+1):
            h += 1
            cell = sheet.cell(m,3).value
            cell1 = sheet.cell(h, 2).value
            cell2 = sheet.cell(h, 4).value
            if cell == None or cell1 == None or cell2 == None:
                continue
            # print('1,',cell, cell1, cell2)
            cell = deal_cell(cell)
            cell1 = deal_cell(cell1)
            cell2 = deal_cell(cell2)
            if cell == d[1]:
                print('计价表,', cell, cell1, cell2)
                print('我的总结', d)
                if cell1 == d[0] and cell2 == d[2]:
                    price = sheet.cell(h,7).value
                    price_list.append(h)
                    price_list.append(price)
                    final_list.append(price_list)
                    print('计价表,', cell, cell1, cell2)
                    print('我的总结', d)
                    break
        if price_list == []:
            price_list = 'error'
            final_list.append(price_list)
    return final_list

def toexcel(pricelist,file,_,path):
    sheet = file.active
    n = sheet.max_column
    for i in range(len(pricelist)):
        # print(pricelist[i])
        if pricelist[i] != 'error':
            sheet.cell(i+3, n + 1, pricelist[i][0])
            sheet.cell(i+3, n + 2, pricelist[i][1])
        else:
            sheet.cell(i + 3, n + 1, '出错')
            sheet.cell(i + 3, n + 2, '出错')
        file.save(filename=path)


def datacheck(data):
    i = 0
    for daa in data:
        for da in daa:
            if da == '':
                data[i] = 'error'
                break
        i += 1
    if data[0] == '应急灯':
        data[1] = "2W~5W×2"
    return data


def sanfang():
    file, _, path = get_path()
    data = getdata(file)
    print(data)
    print('输入基础表地址')
    data = datacheck(data)
    base, a, path1 = get_path()
    x = getprice(base, data, file, _)
    print(x)
    toexcel(x, file, _, path)
    print('写入完成')



